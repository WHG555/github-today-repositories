# !/usr/bin/env python3
# -*- coding: utf-8 -*-
###----------1、文件说明----------###
'''
* 说明：python程序模板
* 时间：
* 文件：
* 作者：Smile
* 版本：0.1
* 备注：
'''
###----------2、库导入----------###
import os, sys
from loguru import logger as log
import pprint
import time
import requests
from lxml import etree
from openpyxl import Workbook
from openpyxl import load_workbook
###----------3、参数配置----------###
###----------4、功能程序----------###

###----------5、主体程序----------###
@log.catch # 捕获报错信息
def getGithubNews(date="day"):
    # 目标网页的URL
    if date == "day":
        url = 'https://github.com/trending'
    elif date == "week":
        url = 'https://github.com/trending?since=weekly'
    elif date == "month":
        url = 'https://github.com/trending?since=monthly'

    # 代理服务器的地址和端口
    proxies = {
        'http': 'http://127.0.0.1:7890',
        'https': 'http://127.0.0.1:7890',
    }

    # 发送HTTP GET请求获取网页内容
    # response = requests.get(url, proxies=proxies)
    response = requests.get(url)

    # 确保请求成功
    if response.status_code != 200:
        with open(date + time.strftime("_%Y-%m-%d_%H-%M-%S", time.localtime()) + ".html" , "w", encoding="utf-8") as f:
            f.write(response.text)
        log.error(response.status_code)
        return

    htmldata = response.text

    # with open("demo.html", "r", encoding="utf-8") as f:
    #     htmldata = f.read()

    # 加载一个已存在的工作簿
    wb = load_workbook('github.xlsx')
    ws = wb[date]

    '''解析网页'''
    tree = etree.HTML(htmldata)
    tab = tree.xpath('//article')
    for dt in tab:
        print("-------------------------")
        print(dt)
        try:
            codelink = dt.xpath(".//*[contains(concat(' ', normalize-space(@class), ' '), ' h3 lh-condensed ')]")[0].xpath(".//a/@href")[0]
        except:
            codelink = ""
        print("代码的路径->", codelink)
        try:
            codereadme = dt.xpath(".//*[contains(concat(' ', normalize-space(@class), ' '), 'col-9 color-fg-muted my-1 pr-4')]/text()")[0].strip()
        except:
            codereadme = ""
        print("说明的文档->", codereadme)
        try:
            codelanguage = dt.xpath('.//span[@itemprop="programmingLanguage"]/text()')[0]
        except:
            codelanguage = ""
        print("编程的语言->", codelanguage)
        try:
            codestar = dt.xpath(".//a[@href='" + codelink + "/stargazers" + "']/text()")[0].strip().replace(",", "")
        except:
            codestar = ""
        print("代码star数->", codestar)
        try:
            codefork = dt.xpath(".//a[@href='" + codelink + "/forks" + "']/text()")[0].strip().replace(",", "")
        except:
            codefork = ""
        print("代码fork数->", codefork)
        try:
            codethisstar = dt.xpath(".//*[contains(concat(' ', normalize-space(@class), ' '), 'd-inline-block float-sm-right')]/text()")[1]
            codethisstar = "".join(''.join([c if c.isdigit() else ' ' for c in codethisstar]).split())
        except:
            codethisstar = ""
        print("今日Star数->", codethisstar)
        ws.append([ codelink.split("/")[-1] ,"https://www.github.com" + codelink, codereadme, codelanguage, codestar, codefork, codethisstar, time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())])

    # 保存工作簿
    wb.save("github.xlsx")

if __name__ == '__main__':
    log.debug('Start Program...')

    if not os.path.exists("github.xlsx"):
        # 创建一个新的工作簿
        wb = Workbook()
        # 创建一个新的工作表
        ws = wb.create_sheet(title="day")
        ws.append(["项目名", "链接", "项目说明", "编程语言", "Star数", "Fork数", "今日增长", "更新时间"])
        ws = wb.create_sheet(title="week")
        ws.append(["项目名", "链接", "项目说明", "编程语言", "Star数", "Fork数", "本周增长", "更新时间"])
        ws = wb.create_sheet(title="month")
        ws.append(["项目名", "链接", "项目说明", "编程语言", "Star数", "Fork数", "当月增长", "更新时间"])
        wb.save("github.xlsx")

    # 添加日志到文件
    log.add("log.log", backtrace=True, diagnose=True, level="DEBUG")

    getGithubNews("day")
    getGithubNews("week")
    getGithubNews("month")
    log.debug('程序运行完成') # 发送系统通知